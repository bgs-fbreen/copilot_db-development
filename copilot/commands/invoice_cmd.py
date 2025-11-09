"""
Invoice management for BGS with precision formatting:
- Table columns: Date, Resource, Task, Sub-Task, Hours, Rate, Total, Description
- Align: Date, Resource, Task, Sub-Task=center; Hours, Rate, Total, totals=right; Description=left
- "na" sub-task is blank
- Subtotal row labeled "SubTotal", right-aligned, bold, size 11
- Invoice Total label/value: both bold, size 14, right aligned
- Headers for invoice details with requested bolds and alignment
"""

import click
from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt, Confirm
from copilot.db import execute_query, get_connection
from datetime import datetime, timedelta
from decimal import Decimal
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, NamedStyle, numbers, Alignment
import re

console = Console()

def sanitize(name):
    if not name:
        return ""
    return re.sub(r'[^a-zA-Z0-9_]', '_', name.strip().replace(' ', '_'))[:32]

def create_project_dir(client_code, project_code, project_abbr):
    PROJECT_BASE = "/mnt/sda1/01_bgm_projman/Active"
    PROJECT_FALLBACK = os.path.expanduser("~/bgm_projects/Active")
    base_dir = PROJECT_BASE if os.path.exists(PROJECT_BASE) else PROJECT_FALLBACK
    project_root = os.path.join(base_dir, client_code, f"{project_code}_{project_abbr}")
    invoice_dir = os.path.join(project_root, '02_invoices')
    Path(invoice_dir).mkdir(parents=True, exist_ok=True)
    return invoice_dir

def clear_screen():
    os.system('clear' if os.name != 'nt' else 'cls')

@click.group()
def invoice():
    """Invoice management for BGS projects"""
    pass

@invoice.command('create')
@click.option('--project', '-p', help='Project code to invoice')
@click.option('--date', '-d', help='Invoice date (YYYY-MM-DD), default: today')
@click.option('--auto', is_flag=True, help='Auto-generate invoice without prompts')
def create_invoice(project, date, auto):
    clear_screen()
    console.print("\n[bold cyan]═══════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   Create Invoice (Draft/Finalize Workflow)[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════[/bold cyan]\n")

    # Get invoice date
    if not date:
        if auto:
            invoice_date = datetime.now().date()
        else:
            date_str = Prompt.ask('Invoice Date', default=datetime.now().strftime('%Y-%m-%d'))
            try:
                invoice_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                console.print(f"[red]Invalid date format: {date_str}[/red]")
                return
    else:
        try:
            invoice_date = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            console.print(f"[red]Invalid date format: {date}[/red]")
            return

    # Get projects with unbilled time
    if project:
        projects = execute_query("""
            SELECT 
                p.project_code, p.project_name, p.client_code, c.name as client_name, p.client_po,
                COUNT(t.id) as unbilled_count,
                SUM(t.ts_units) as total_hours,
                SUM(t.ts_mileage) as total_miles,
                SUM(t.ts_expense) as total_expenses
            FROM bgs.project p
            JOIN bgs.client c ON c.code = p.client_code
            JOIN bgs.timesheet t ON t.project_code = p.project_code
            WHERE t.invoice_code IS NULL
              AND p.project_code = %s
            GROUP BY p.project_code, p.project_name, p.client_code, c.name, p.client_po
        """, (project,))
    else:
        projects = execute_query("""
            SELECT 
                p.project_code, p.project_name, p.client_code, c.name as client_name, p.client_po,
                COUNT(t.id) as unbilled_count,
                SUM(t.ts_units) as total_hours,
                SUM(t.ts_mileage) as total_miles,
                SUM(t.ts_expense) as total_expenses
            FROM bgs.project p
            JOIN bgs.client c ON c.code = p.client_code
            JOIN bgs.timesheet t ON t.project_code = p.project_code
            WHERE t.invoice_code IS NULL
            GROUP BY p.project_code, p.project_name, p.client_code, c.name, p.client_po
            ORDER BY p.project_code
        """)

    if not projects:
        console.print("[yellow]No unbilled time entries found.[/yellow]")
        if project:
            console.print(f"[yellow]Project: {project}[/yellow]")
        return

    # Display projects with unbilled time
    console.print("[bold]Projects with Unbilled Time:[/bold]\n")
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("Project", style="cyan")
    table.add_column("Client", style="green")
    table.add_column("Project Name", style="white")
    table.add_column("Entries", justify="right", style="yellow")
    table.add_column("Hours", justify="right")
    table.add_column("Miles", justify="right")
    table.add_column("Expenses", justify="right")
    for proj in projects:
        table.add_row(
            proj['project_code'],
            proj['client_code'],
            (proj['project_name'] or '')[:40],
            str(proj['unbilled_count']),
            f"{float(proj['total_hours'] or 0):.1f}",
            f"{float(proj['total_miles'] or 0):.1f}",
            f"${float(proj['total_expenses'] or 0):.2f}"
        )
    console.print(table)
    console.print()

    # Prompt for project if not provided
    if not project:
        if auto:
            console.print("[red]Error: --project required with --auto[/red]")
            return
        project = Prompt.ask('Project Code to invoice')
    proj_data = [p for p in projects if p['project_code'] == project]
    if not proj_data:
        console.print(f"[red]Project '{project}' not found or has no unbilled time[/red]")
        return
    proj = proj_data[0]

    timesheets = execute_query("""
        SELECT 
            t.id, t.ts_date, t.task_no, t.sub_task_no, t.res_id, t.ts_units,
            t.ts_desc, b.base_rate, b.is_lump_sum, t.ts_expense
        FROM bgs.timesheet t
        LEFT JOIN bgs.baseline b ON 
            b.project_code = t.project_code AND b.task_no = t.task_no AND b.sub_task_no = t.sub_task_no AND b.res_id = t.res_id
        WHERE t.project_code = %s AND t.invoice_code IS NULL
        ORDER BY t.ts_date, t.task_no, t.sub_task_no
    """, (project,))
    if not timesheets:
        console.print(f"[yellow]No unbilled timesheets for project {project}[/yellow]")
        return

    labor_table = []
    total_hours = Decimal('0')
    subtotal = Decimal('0')
    for ts in timesheets:
        is_lump_sum = bool(ts.get('is_lump_sum'))
        hours = Decimal(str(ts['ts_units'] or 0))
        rate = Decimal(str(ts['base_rate'] or 0))
        expense = Decimal(str(ts['ts_expense'] or 0))
        subtask_val = ts['sub_task_no']
        if subtask_val is not None and str(subtask_val).lower() == "na":
            subtask_val = ""
        if is_lump_sum:
            labor_table.append({
                'date': ts['ts_date'],
                'resource': ts['res_id'],
                'task': ts['task_no'],
                'subtask': subtask_val,
                'hours': "",
                'rate': "",
                'total': expense,
                'description': ts['ts_desc'] or ''
            })
            subtotal += expense
        else:
            total = hours * rate
            labor_table.append({
                'date': ts['ts_date'],
                'resource': ts['res_id'],
                'task': ts['task_no'],
                'subtask': subtask_val,
                'hours': hours,
                'rate': rate,
                'total': total,
                'description': ts['ts_desc'] or ''
            })
            total_hours += hours
            subtotal += total

    invoice_total = subtotal

    # XLSX DRAFT (pre-finalize)
    invoice_number = None
    project_abbr = sanitize(proj['project_name'] if proj['project_name'] else proj['project_code'])
    invoice_dir = create_project_dir(proj['client_code'], project, project_abbr)
    year = invoice_date.year
    last_invoice = execute_query(
        "SELECT MAX(invoice_number) as last_num FROM bgs.invoice WHERE project_code = %s AND EXTRACT(YEAR FROM invoice_date) = %s",
        (project, year)
    )
    if last_invoice and last_invoice[0]['last_num']:
        invoice_number = last_invoice[0]['last_num'] + 1
    else:
        invoice_number = 1
    invoice_code = f"{project}.{invoice_number:04d}"
    xlsx_file = os.path.join(invoice_dir, f"{invoice_code}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Invoice Draft"

    ws['F1'] = "Invoice"
    ws['F1'].font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 28

    ws['A2'] = "Breen GeoScience Management, Inc."
    ws['A2'].font = Font(bold=True)
    ws['A3'] = "Payment to:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = "Breen GeoScience Management, Inc."
    ws['A5'] = "PMB #354, 4234 I-75 Business Spur"
    ws['A6'] = "Sault Ste. Marie, Michigan USA 49783"
    ws['F2'] = "Project"
    ws['F2'].font = Font(bold=True)
    ws['F3'] = proj['project_name'] or ""
    ws['F4'] = ""
    ws['F5'] = "USA"
    ws['F7'] = f"Date: {invoice_date.strftime('%B %d, %Y')}"
    ws['A8'] = "Client"
    ws['A8'].font = Font(bold=True)
    ws['A9'] = proj['client_name']
    ws['A10'] = ""
    ws['A11'] = ""
    ws['A12'] = ""
    ws['A13'] = ""
    ws['A14'] = ""
    ws['F8'] = f"Desc. {project}"
    ws['F9'] = f"PO#: {proj['client_po'] or ''}"
    ws['F10'] = "Net 30"
    ws['F11'] = f"Baseline: {project}"
    ws['F13'] = f"Invoice No. {invoice_code}"

    currency_fmt = NamedStyle(name="currency_fmt", number_format=numbers.FORMAT_CURRENCY_USD_SIMPLE)
    int_fmt = NamedStyle(name="int_fmt", number_format='#,##0')
    if "currency_fmt" not in wb.named_styles:
        wb.add_named_style(currency_fmt)
    if "int_fmt" not in wb.named_styles:
        wb.add_named_style(int_fmt)

    # Labor Table header with alignment
    cur_row = 15
    headers = ["Date", "Resource", "Task", "Sub-Task", "Hours", "Rate", "Total", "Description"]
    aligns =   ["center", "center", "center", "center",   "right", "right", "right", "left"]
    for idx, (text, halign) in enumerate(zip(headers, aligns), 1):
        ws.cell(row=cur_row, column=idx, value=text)
        ws.cell(row=cur_row, column=idx).font = Font(bold=True)
        ws.cell(row=cur_row, column=idx).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.cell(row=cur_row, column=idx).alignment = Alignment(horizontal=halign)
    cur_row += 1

    for item in labor_table:
        ws[f"A{cur_row}"] = item['date'].strftime('%Y-%m-%d')
        ws[f"A{cur_row}"].alignment = Alignment(horizontal="center")
        ws[f"B{cur_row}"] = item['resource']
        ws[f"B{cur_row}"].alignment = Alignment(horizontal="center")
        ws[f"C{cur_row}"] = item['task']
        ws[f"C{cur_row}"].alignment = Alignment(horizontal="center")
        ws[f"D{cur_row}"] = item['subtask'] if item['subtask'] is not None else ""
        ws[f"D{cur_row}"].alignment = Alignment(horizontal="center")
        ws[f"E{cur_row}"] = float(item['hours']) if item['hours'] not in ("", None) else ""
        ws[f"E{cur_row}"].alignment = Alignment(horizontal="right")
        if item['hours'] not in ("", None):
            ws[f"E{cur_row}"].style = "int_fmt"
        ws[f"F{cur_row}"] = float(item['rate']) if item['rate'] not in ("", None) else ""
        ws[f"F{cur_row}"].alignment = Alignment(horizontal="right")
        if item['rate'] not in ("", None):
            ws[f"F{cur_row}"].style = "currency_fmt"
        ws[f"G{cur_row}"] = float(item['total']) if item['total'] is not None else ""
        ws[f"G{cur_row}"].alignment = Alignment(horizontal="right")
        if item['total'] is not None:
            ws[f"G{cur_row}"].style = "currency_fmt"
        ws[f"H{cur_row}"] = item['description']
        ws[f"H{cur_row}"].alignment = Alignment(horizontal="left")
        cur_row += 1

    # Subtotal row (right-aligned, bold, size 11)
    ws[f"F{cur_row}"] = "SubTotal"
    ws[f"F{cur_row}"].alignment = Alignment(horizontal="right")
    ws[f"F{cur_row}"].font = Font(bold=True)
    ws[f"G{cur_row}"] = float(subtotal)
    ws[f"G{cur_row}"].alignment = Alignment(horizontal="right")
    ws[f"G{cur_row}"].font = Font(bold=True)
    ws[f"G{cur_row}"].style = "currency_fmt"
    cur_row += 1

    # Invoice Total label/value, both bold and size 14, right aligned
    ws[f"F{cur_row}"] = "Invoice Total:"
    ws[f"F{cur_row}"].alignment = Alignment(horizontal="right")
    ws[f"F{cur_row}"].font = Font(bold=True, size=14)
    ws[f"G{cur_row}"] = float(invoice_total)
    ws[f"G{cur_row}"].alignment = Alignment(horizontal="right")
    ws[f"G{cur_row}"].font = Font(bold=True, size=14)
    ws[f"G{cur_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.row_dimensions[cur_row].height = 28  # Increase row height for the larger font
    for col in 'ABCDEFGH':
        ws.column_dimensions[col].width = 17 if col != 'H' else 36
    wb.save(xlsx_file)

    console.print(f"[green]✓ Invoice DRAFT written as: {xlsx_file}[/green]\n")
    console.print("[bold yellow]This invoice is in DRAFT. No DB update yet.[/bold yellow]\n")

    # 8. Prompt to finalize (commit)
    if Confirm.ask(f"Do you want to finalize and record invoice {invoice_code} (DB will be written)?", default=False):
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO bgs.invoice
                    (invoice_code, project_code, invoice_number, invoice_date, 
                     due_date, amount, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (invoice_code, project, invoice_number, invoice_date,
                      invoice_date + timedelta(days=30), invoice_total, 'pending'))
                timesheet_ids = [ts['id'] for ts in timesheets]
                cur.execute("""
                    UPDATE bgs.timesheet
                    SET invoice_code = %s
                    WHERE id = ANY(%s)
                """, (invoice_code, timesheet_ids))
                conn.commit()
            console.print(f"[green]✓ Invoice {invoice_code} finalized and recorded![/green]\n")
        except Exception as e:
            conn.rollback()
            console.print(f"[red]Error finalizing invoice: {e}[/red]")
        finally:
            conn.close()
    else:
        console.print(f"[yellow]Invoice {invoice_code} left as DRAFT. You may edit/re-run as needed.[/yellow]\n")
