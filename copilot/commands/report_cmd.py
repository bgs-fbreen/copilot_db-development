"""
Financial reports and analytics for Copilot Accounting System
"""
import click
from rich.console import Console
from rich.table import Table
from copilot.db import execute_query
from datetime import datetime, date, timedelta
import csv
import os
from pathlib import Path

console = Console()

def clear_screen():
    """Clear the terminal screen"""
    os.system('clear' if os.name != 'nt' else 'cls')

def get_report_dir():
    """Get or create the reports directory"""
    report_dir = os.path.expanduser("~/copilot_reports")
    Path(report_dir).mkdir(parents=True, exist_ok=True)
    return report_dir

def export_to_csv(data, filename, headers):
    """Export data to CSV file"""
    report_dir = get_report_dir()
    filepath = os.path.join(report_dir, filename)
    
    with open(filepath, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        for row in data:
            writer.writerow(row)
    
    return filepath

def export_to_xlsx(data, filename, headers, title):
    """Export data to XLSX file using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        report_dir = get_report_dir()
        filepath = os.path.join(report_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row_idx, row_data in enumerate(data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        return filepath
    except ImportError:
        console.print("[yellow]openpyxl not installed[/yellow]")
        return None

@click.group()
def report():
    """Financial reports and analytics"""
    pass

@report.command('pl')
@click.option('--entity', '-e', help='Filter by entity (BGS, MHB)')
@click.option('--year', '-y', type=int, help='Filter by year')
@click.option('--month', '-m', type=int, help='Filter by month (1-12)')
@click.option('--quarter', '-q', type=click.Choice(['Q1', 'Q2', 'Q3', 'Q4']), help='Filter by quarter')
@click.option('--export', type=click.Choice(['csv', 'xlsx']), help='Export format')
def profit_loss(entity, year, month, quarter, export):
    """Generate Profit & Loss statement"""
    clear_screen()
    
    if not year:
        year = date.today().year
    
    # Build filter conditions
    conditions = ["EXTRACT(YEAR FROM t.trans_date) = %s"]
    params = [year]
    
    if entity:
        conditions.append("t.entity = %s")
        params.append(entity)
    
    if month:
        conditions.append("EXTRACT(MONTH FROM t.trans_date) = %s")
        params.append(month)
    
    if quarter:
        quarter_num = int(quarter[1])
        start_month = (quarter_num - 1) * 3 + 1
        end_month = start_month + 2
        conditions.append("EXTRACT(MONTH FROM t.trans_date) BETWEEN %s AND %s")
        params.extend([start_month, end_month])
    
    where_clause = " AND ".join(conditions)
    
    # Query for income
    income_query = f"""
        SELECT c.name, SUM(t.amount) as total
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE c.account_type = 'income'
          AND {where_clause}
        GROUP BY c.name
        ORDER BY total DESC
    """
    income_data = execute_query(income_query, params)
    
    # Query for expenses
    expense_query = f"""
        SELECT c.name, SUM(ABS(t.amount)) as total
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE c.account_type = 'expense'
          AND {where_clause}
        GROUP BY c.name
        ORDER BY total DESC
    """
    expense_data = execute_query(expense_query, params)
    
    # Calculate totals
    total_income = sum(float(row['total']) for row in income_data) if income_data else 0
    total_expenses = sum(float(row['total']) for row in expense_data) if expense_data else 0
    net_income = total_income - total_expenses
    
    # Prepare title
    period_str = ""
    if month:
        period_str = f"{date(year, month, 1).strftime('%B')} {year}"
    elif quarter:
        period_str = f"{quarter} {year}"
    else:
        period_str = str(year)
    
    entity_str = f"{entity} - " if entity else ""
    
    # Display report
    console.print("\n[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   Profit & Loss Statement[/bold cyan]")
    console.print(f"[bold cyan]   {entity_str}{period_str}[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    console.print("[bold]INCOME[/bold]")
    for row in income_data:
        console.print(f"   {row['name']:<35} ${float(row['total']):>12,.2f}")
    console.print("   " + "─" * 52)
    console.print(f"   [bold]Total Income{' ' * 23} ${total_income:>12,.2f}[/bold]\n")
    
    console.print("[bold]EXPENSES[/bold]")
    for row in expense_data:
        console.print(f"   {row['name']:<35} ${float(row['total']):>12,.2f}")
    console.print("   " + "─" * 52)
    console.print(f"   [bold]Total Expenses{' ' * 21} ${total_expenses:>12,.2f}[/bold]\n")
    
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    net_color = "green" if net_income >= 0 else "red"
    console.print(f"   [bold {net_color}]NET INCOME{' ' * 25} ${net_income:>12,.2f}[/bold {net_color}]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    # Export if requested
    if export:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        entity_suffix = f"_{entity}" if entity else ""
        filename = f"pl{entity_suffix}_{year}"
        if month:
            filename += f"_{month:02d}"
        elif quarter:
            filename += f"_{quarter}"
        filename += f"_{timestamp}"
        
        # Prepare export data
        export_data = []
        export_data.append(["INCOME", "", ""])
        for row in income_data:
            export_data.append(["", row['name'], float(row['total'])])
        export_data.append(["", "Total Income", total_income])
        export_data.append(["", "", ""])
        export_data.append(["EXPENSES", "", ""])
        for row in expense_data:
            export_data.append(["", row['name'], float(row['total'])])
        export_data.append(["", "Total Expenses", total_expenses])
        export_data.append(["", "", ""])
        export_data.append(["NET INCOME", "", net_income])
        
        if export == 'csv':
            filepath = export_to_csv(export_data, f"{filename}.csv", ["Section", "Category", "Amount"])
            console.print(f"[green]Report exported to: {filepath}[/green]\n")
        elif export == 'xlsx':
            title = f"Profit & Loss - {entity_str}{period_str}"
            filepath = export_to_xlsx(export_data, f"{filename}.xlsx", ["Section", "Category", "Amount"], title)
            if filepath:
                console.print(f"[green]Report exported to: {filepath}[/green]\n")

@report.command('cashflow')
@click.option('--year', '-y', type=int, help='Filter by year')
@click.option('--month', '-m', type=int, help='Filter by month (1-12)')
@click.option('--entity', '-e', help='Filter by entity (BGS, MHB)')
def cashflow(year, month, entity):
    """Generate Cash Flow statement"""
    clear_screen()
    
    if not year:
        year = date.today().year
    
    # Build filter conditions
    conditions = ["EXTRACT(YEAR FROM t.trans_date) = %s"]
    params = [year]
    
    if month:
        conditions.append("EXTRACT(MONTH FROM t.trans_date) = %s")
        params.append(month)
    
    if entity:
        conditions.append("t.entity = %s")
        params.append(entity)
    
    where_clause = " AND ".join(conditions)
    
    # Get opening balance
    opening_query = "SELECT COALESCE(SUM(t.amount), 0) as balance FROM acc.transaction t WHERE t.trans_date < %s"
    opening_params = [date(year, month if month else 1, 1)]
    if entity:
        opening_query += " AND t.entity = %s"
        opening_params.append(entity)
    
    opening_result = execute_query(opening_query, opening_params)
    opening_balance = float(opening_result[0]['balance']) if opening_result else 0
    
    # Get cash inflows
    inflows_query = f"""
        SELECT c.name, SUM(t.amount) as total
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE {where_clause} AND t.amount > 0
        GROUP BY c.name
        ORDER BY total DESC
    """
    inflows = execute_query(inflows_query, params)
    
    # Get cash outflows
    outflows_query = f"""
        SELECT c.name, SUM(ABS(t.amount)) as total
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE {where_clause} AND t.amount < 0
        GROUP BY c.name
        ORDER BY total DESC
    """
    outflows = execute_query(outflows_query, params)
    
    # Calculate totals
    total_inflows = sum(float(row['total']) for row in inflows) if inflows else 0
    total_outflows = sum(float(row['total']) for row in outflows) if outflows else 0
    net_cashflow = total_inflows - total_outflows
    closing_balance = opening_balance + net_cashflow
    
    # Display report
    period_str = f"{date(year, month, 1).strftime('%B')} {year}" if month else str(year)
    entity_str = f" - {entity}" if entity else ""
    
    console.print("\n[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   Cash Flow Statement[/bold cyan]")
    console.print(f"[bold cyan]   {period_str}{entity_str}[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    console.print(f"[bold]Opening Balance{' ' * 21} ${opening_balance:>12,.2f}[/bold]\n")
    
    console.print("[bold]CASH INFLOWS[/bold]")
    for row in inflows:
        console.print(f"   {row['name']:<35} ${float(row['total']):>12,.2f}")
    console.print("   " + "─" * 52)
    console.print(f"   [bold]Total Inflows{' ' * 22} ${total_inflows:>12,.2f}[/bold]\n")
    
    console.print("[bold]CASH OUTFLOWS[/bold]")
    for row in outflows:
        console.print(f"   {row['name']:<35} ${float(row['total']):>12,.2f}")
    console.print("   " + "─" * 52)
    console.print(f"   [bold]Total Outflows{' ' * 21} ${total_outflows:>12,.2f}[/bold]\n")
    
    net_color = "green" if net_cashflow >= 0 else "red"
    console.print(f"[{net_color}]Net Cash Flow{' ' * 23} ${net_cashflow:>12,.2f}[/{net_color}]\n")
    
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    console.print(f"   [bold]Closing Balance{' ' * 21} ${closing_balance:>12,.2f}[/bold]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")

@report.command('tax')
@click.option('--year', '-y', type=int, help='Filter by year')
@click.option('--export', type=click.Choice(['csv', 'xlsx']), help='Export format')
def tax(year, export):
    """Generate Tax Summary report"""
    clear_screen()
    
    if not year:
        year = date.today().year
    
    # Get taxable income
    income_query = """
        SELECT c.name, SUM(t.amount) as total
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE c.account_type = 'income' AND c.is_taxable = true
          AND EXTRACT(YEAR FROM t.trans_date) = %s
        GROUP BY c.name ORDER BY total DESC
    """
    taxable_income = execute_query(income_query, [year])
    
    # Get deductible expenses
    deductible_query = """
        SELECT c.name, SUM(ABS(t.amount)) as total
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE c.account_type = 'expense' AND c.is_taxable = true
          AND EXTRACT(YEAR FROM t.trans_date) = %s
        GROUP BY c.name ORDER BY total DESC
    """
    deductible_expenses = execute_query(deductible_query, [year])
    
    # Get non-deductible expenses
    non_deductible_query = """
        SELECT c.name, SUM(ABS(t.amount)) as total
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE c.account_type = 'expense' AND c.is_taxable = false
          AND EXTRACT(YEAR FROM t.trans_date) = %s
        GROUP BY c.name ORDER BY total DESC
    """
    non_deductible = execute_query(non_deductible_query, [year])
    
    # Calculate totals
    total_taxable_income = sum(float(row['total']) for row in taxable_income) if taxable_income else 0
    total_deductible = sum(float(row['total']) for row in deductible_expenses) if deductible_expenses else 0
    total_non_deductible = sum(float(row['total']) for row in non_deductible) if non_deductible else 0
    net_taxable_income = total_taxable_income - total_deductible
    
    # Display report
    console.print("\n[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   Tax Summary Report[/bold cyan]")
    console.print(f"[bold cyan]   {year}[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    console.print("[bold]TAXABLE INCOME[/bold]")
    for row in taxable_income:
        console.print(f"   {row['name']:<35} ${float(row['total']):>12,.2f}")
    console.print("   " + "─" * 52)
    console.print(f"   [bold]Total Taxable Income{' ' * 15} ${total_taxable_income:>12,.2f}[/bold]\n")
    
    console.print("[bold]DEDUCTIBLE EXPENSES[/bold]")
    for row in deductible_expenses:
        console.print(f"   {row['name']:<35} ${float(row['total']):>12,.2f}")
    console.print("   " + "─" * 52)
    console.print(f"   [bold]Total Deductible{' ' * 21} ${total_deductible:>12,.2f}[/bold]\n")
    
    if non_deductible:
        console.print("[bold]NON-DEDUCTIBLE EXPENSES[/bold]")
        for row in non_deductible:
            console.print(f"   {row['name']:<35} ${float(row['total']):>12,.2f}")
        console.print("   " + "─" * 52)
        console.print(f"   [bold]Total Non-Deductible{' ' * 16} ${total_non_deductible:>12,.2f}[/bold]\n")
    
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    net_color = "green" if net_taxable_income >= 0 else "red"
    console.print(f"   [bold {net_color}]NET TAXABLE INCOME{' ' * 17} ${net_taxable_income:>12,.2f}[/bold {net_color}]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    # Export if requested
    if export:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"tax_{year}_{timestamp}"
        export_data = []
        export_data.append(["TAXABLE INCOME", "", ""])
        for row in taxable_income:
            export_data.append(["", row['name'], float(row['total'])])
        export_data.append(["", "Total Taxable Income", total_taxable_income])
        export_data.append(["", "", ""])
        export_data.append(["DEDUCTIBLE EXPENSES", "", ""])
        for row in deductible_expenses:
            export_data.append(["", row['name'], float(row['total'])])
        export_data.append(["", "Total Deductible", total_deductible])
        if non_deductible:
            export_data.append(["", "", ""])
            export_data.append(["NON-DEDUCTIBLE EXPENSES", "", ""])
            for row in non_deductible:
                export_data.append(["", row['name'], float(row['total'])])
            export_data.append(["", "Total Non-Deductible", total_non_deductible])
        export_data.append(["", "", ""])
        export_data.append(["NET TAXABLE INCOME", "", net_taxable_income])
        
        if export == 'csv':
            filepath = export_to_csv(export_data, f"{filename}.csv", ["Section", "Category", "Amount"])
            console.print(f"[green]Report exported to: {filepath}[/green]\n")
        elif export == 'xlsx':
            title = f"Tax Summary - {year}"
            filepath = export_to_xlsx(export_data, f"{filename}.xlsx", ["Section", "Category", "Amount"], title)
            if filepath:
                console.print(f"[green]Report exported to: {filepath}[/green]\n")

@report.command('monthly')
@click.option('--year', '-y', type=int, help='Filter by year')
@click.option('--entity', '-e', help='Filter by entity (BGS, MHB)')
def monthly(year, entity):
    """Generate Monthly Summary report"""
    clear_screen()
    
    if not year:
        year = date.today().year
    
    # Build filter conditions
    conditions = ["EXTRACT(YEAR FROM t.trans_date) = %s"]
    params = [year]
    
    if entity:
        conditions.append("t.entity = %s")
        params.append(entity)
    
    where_clause = " AND ".join(conditions)
    
    # Query for monthly summary
    summary_query = f"""
        SELECT 
            EXTRACT(MONTH FROM t.trans_date) as month,
            SUM(CASE WHEN c.account_type = 'income' THEN t.amount ELSE 0 END) as income,
            SUM(CASE WHEN c.account_type = 'expense' THEN ABS(t.amount) ELSE 0 END) as expenses
        FROM acc.transaction t
        JOIN acc.category c ON c.id = t.category_id
        WHERE {where_clause}
        GROUP BY EXTRACT(MONTH FROM t.trans_date)
        ORDER BY month
    """
    monthly_data = execute_query(summary_query, params)
    
    # Create monthly dictionary
    months_dict = {int(row['month']): row for row in monthly_data}
    
    # Display report
    entity_str = f" - {entity}" if entity else ""
    console.print("\n[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   Monthly Summary Report[/bold cyan]")
    console.print(f"[bold cyan]   {year}{entity_str}[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    # Create table
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("Month", style="cyan")
    table.add_column("Income", justify="right", style="green")
    table.add_column("Expenses", justify="right", style="red")
    table.add_column("Net", justify="right", style="bold white")
    table.add_column("YTD Net", justify="right", style="bold yellow")
    
    ytd_net = 0
    for month_num in range(1, 13):
        month_name = date(year, month_num, 1).strftime('%b %Y')
        
        if month_num in months_dict:
            income = float(months_dict[month_num]['income'] or 0)
            expenses = float(months_dict[month_num]['expenses'] or 0)
            net = income - expenses
            ytd_net += net
            
            net_color = "green" if net >= 0 else "red"
            ytd_color = "green" if ytd_net >= 0 else "red"
            
            table.add_row(
                month_name,
                f"${income:,.2f}",
                f"${expenses:,.2f}",
                f"[{net_color}]${net:,.2f}[/{net_color}]",
                f"[{ytd_color}]${ytd_net:,.2f}[/{ytd_color}]"
            )
        else:
            table.add_row(month_name, "$0.00", "$0.00", "$0.00", f"${ytd_net:,.2f}")
    
    console.print(table)
    console.print()

@report.command('project')
@click.argument('project_code', required=False)
@click.option('--all', is_flag=True, help='Show all active projects')
def project_report(project_code, all):
    """Generate Project Profitability report"""
    clear_screen()
    
    if not project_code and not all:
        console.print("[red]Error: Provide a project code or use --all flag[/red]\n")
        return
    
    # Build query
    if all:
        where_clause = "p.status = 'active'"
        params = []
    else:
        where_clause = "p.project_code = %s"
        params = [project_code]
    
    # Get project data
    project_query = f"""
        SELECT 
            p.project_code, p.project_name, p.client_code, c.name as client_name, p.status,
            COALESCE(SUM(i.amount), 0) as total_invoiced,
            COALESCE(SUM(i.paid_amount), 0) as total_paid
        FROM bgs.project p
        JOIN bgs.client c ON c.code = p.client_code
        LEFT JOIN bgs.invoice i ON i.project_code = p.project_code
        WHERE {where_clause}
        GROUP BY p.project_code, p.project_name, p.client_code, c.name, p.status
        ORDER BY p.project_code
    """
    projects = execute_query(project_query, params)
    
    if not projects:
        console.print("[yellow]No projects found[/yellow]\n")
        return
    
    console.print("\n[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   Project Profitability Report[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    for proj in projects:
        # Get labor and mileage costs in a single query
        costs_query = """
            SELECT 
                COALESCE(SUM(t.ts_units * COALESCE(b.base_rate, 0)), 0) as labor_cost,
                COALESCE(SUM(t.ts_mileage * COALESCE(b.base_miles_rate, 0)), 0) as mileage_cost,
                COALESCE(SUM(t.ts_expense), 0) as expense_cost
            FROM bgs.timesheet t
            LEFT JOIN bgs.baseline b ON 
                b.project_code = t.project_code 
                AND b.task_no = t.task_no 
                AND b.sub_task_no = t.sub_task_no
                AND b.res_id = t.res_id
            WHERE t.project_code = %s
        """
        costs_result = execute_query(costs_query, [proj['project_code']])
        labor_cost = float(costs_result[0]['labor_cost']) if costs_result else 0
        mileage_cost = float(costs_result[0]['mileage_cost']) if costs_result else 0
        expense_cost = float(costs_result[0]['expense_cost']) if costs_result else 0
        
        total_cost = labor_cost + mileage_cost + expense_cost
        total_invoiced = float(proj['total_invoiced'] or 0)
        total_paid = float(proj['total_paid'] or 0)
        outstanding_ar = total_invoiced - total_paid
        
        # Accrual-based profit (invoiced revenue vs costs)
        gross_profit = total_invoiced - total_cost
        gross_margin = (gross_profit / total_invoiced * 100) if total_invoiced > 0 else 0
        
        # Cash-based position (actual cash received vs costs spent)
        cash_position = total_paid - total_cost
        
        # Display project
        console.print(f"[bold cyan]Project:[/bold cyan] {proj['project_code']}")
        console.print(f"[bold]Client:[/bold] {proj['client_name']} ({proj['client_code']})")
        console.print(f"[bold]Name:[/bold] {proj['project_name']}")
        console.print(f"[bold]Status:[/bold] {proj['status']}\n")
        
        # Revenue section
        console.print(f"[bold]Revenue:[/bold]")
        console.print(f"  Total Invoiced:      ${total_invoiced:>12,.2f}")
        console.print(f"  Total Paid:          ${total_paid:>12,.2f}")
        console.print(f"  Outstanding AR:      ${outstanding_ar:>12,.2f}\n")
        
        # Costs section
        console.print(f"[bold]Costs:[/bold]")
        console.print(f"  Labor:               ${labor_cost:>12,.2f}")
        console.print(f"  Mileage:             ${mileage_cost:>12,.2f}")
        console.print(f"  Direct Expenses:     ${expense_cost:>12,.2f}")
        console.print(f"  Total Costs:         ${total_cost:>12,.2f}\n")
        
        # Profitability section
        console.print(f"[bold]Profitability:[/bold]")
        gross_color = "green" if gross_profit >= 0 else "red"
        gross_label = "Gross Profit:" if gross_profit >= 0 else "Gross Loss:"
        console.print(f"  [{gross_color}]{gross_label:<18} ${abs(gross_profit):>12,.2f}[/{gross_color}]")
        console.print(f"  [{gross_color}]Gross Margin:      {gross_margin:>12.1f}%[/{gross_color}]")
        
        # Cash position (shows actual cash flow status)
        cash_color = "green" if cash_position >= 0 else "yellow"
        cash_label = "Cash Position:" if cash_position >= 0 else "Cash Deficit:"
        console.print(f"  [{cash_color}]{cash_label:<18} ${abs(cash_position):>12,.2f}[/{cash_color}]")
        if cash_position < 0 and outstanding_ar > 0:
            console.print(f"  [dim](Awaiting ${outstanding_ar:,.2f} in AR)[/dim]")
        
        if all:
            console.print("\n" + "─" * 60 + "\n")
    
    console.print()

@report.command('ar')
@click.option('--all', is_flag=True, help='Show all invoices including paid')
@click.option('--export', type=click.Choice(['csv', 'xlsx']), help='Export format')
def ar_export(all, export):
    """Accounts Receivable aging report with export"""
    from copilot.commands.ar_cmd import ar as ar_command
    
    if not export:
        # Just run the regular AR command
        import sys
        from click.testing import CliRunner
        runner = CliRunner()
        runner.invoke(ar_command, ['--all'] if all else [])
        return
    
    # Query for outstanding invoices
    status_filter = "" if all else "WHERE i.status = 'pending'"
    
    invoices = execute_query(f"""
        SELECT 
            i.invoice_code, i.project_code, c.code as client_code, c.name as client_name,
            i.invoice_number, i.invoice_date, i.due_date, i.amount, i.paid_amount, i.status
        FROM bgs.invoice i
        JOIN bgs.project p ON p.project_code = i.project_code
        JOIN bgs.client c ON c.code = p.client_code
        {status_filter}
        ORDER BY i.project_code ASC, i.invoice_number ASC
    """)
    
    if not invoices:
        console.print("[yellow]No outstanding invoices found[/yellow]\n")
        return
    
    # Prepare export data
    today = date.today()
    export_data = []
    
    for inv in invoices:
        invoice_date = inv['invoice_date']
        due_date = inv['due_date'] if inv['due_date'] else invoice_date + timedelta(days=30)
        days_outstanding = (today - invoice_date).days
        
        amount = float(inv['amount'] or 0)
        paid = float(inv['paid_amount'] or 0)
        balance = amount - paid
        
        # Categorize by aging
        current = balance if days_outstanding < 30 else 0
        days_30_60 = balance if 30 <= days_outstanding < 61 else 0
        days_61_90 = balance if 61 <= days_outstanding < 91 else 0
        over_90 = balance if days_outstanding >= 91 else 0
        
        export_data.append([
            inv['client_code'], inv['project_code'], inv['invoice_number'],
            invoice_date.strftime('%Y-%m-%d'), due_date.strftime('%Y-%m-%d'),
            days_outstanding, current, days_30_60, days_61_90, over_90, balance
        ])
    
    # Export
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ar_aging_{timestamp}"
    headers = ["Client", "Project", "Inv#", "Invoice Date", "Due Date", "Days", "Current", "30-60", "61-90", ">90", "Total"]
    
    if export == 'csv':
        filepath = export_to_csv(export_data, f"{filename}.csv", headers)
        console.print(f"\n[green]AR report exported to: {filepath}[/green]\n")
    elif export == 'xlsx':
        title = f"AR Aging Report - {today.strftime('%Y-%m-%d')}"
        filepath = export_to_xlsx(export_data, f"{filename}.xlsx", headers, title)
        if filepath:
            console.print(f"\n[green]AR report exported to: {filepath}[/green]\n")

@report.command('property')
@click.argument('property_code', required=False)
@click.option('--all', is_flag=True, help='Show all properties')
@click.option('--year', '-y', type=int, help='Filter by year')
def property_report(property_code, all, year):
    """Generate Property Report (MHB)"""
    clear_screen()
    
    if not property_code and not all:
        console.print("[red]Error: Provide a property code or use --all flag[/red]\n")
        return
    
    if not year:
        year = date.today().year
    
    # Build query
    if all:
        where_clause = "p.code IS NOT NULL"
        params = []
    else:
        where_clause = "p.code = %s"
        params = [property_code]
    
    # Get property data
    property_query = f"""
        SELECT DISTINCT p.code, p.address, p.city, p.state
        FROM mhb.property p
        WHERE {where_clause}
        ORDER BY p.code
    """
    properties = execute_query(property_query, params)
    
    if not properties:
        console.print("[yellow]No properties found[/yellow]\n")
        return
    
    console.print("\n[bold cyan]═══════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   Property Report (MHB)[/bold cyan]")
    console.print(f"[bold cyan]   {year}[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════[/bold cyan]\n")
    
    for prop in properties:
        # Get rent collected
        rent_query = """
            SELECT COALESCE(SUM(amount), 0) as rent_collected
            FROM mhb.rent_payment
            WHERE property_code = %s AND EXTRACT(YEAR FROM payment_date) = %s
        """
        rent_result = execute_query(rent_query, [prop['code'], year])
        rent_collected = float(rent_result[0]['rent_collected']) if rent_result else 0
        
        # Get expenses
        expense_query = """
            SELECT c.name, SUM(ABS(t.amount)) as total
            FROM acc.transaction t
            JOIN acc.category c ON c.id = t.category_id
            WHERE t.property_code = %s AND EXTRACT(YEAR FROM t.trans_date) = %s AND t.amount < 0
            GROUP BY c.name
            ORDER BY total DESC
        """
        expenses = execute_query(expense_query, [prop['code'], year])
        
        total_expenses = sum(float(row['total']) for row in expenses) if expenses else 0
        net_income = rent_collected - total_expenses
        
        # Display property
        console.print(f"[bold cyan]Property:[/bold cyan] {prop['code']}")
        console.print(f"[bold]Address:[/bold] {prop['address']}, {prop['city']}, {prop['state']}\n")
        
        console.print(f"Rent Collected:        ${rent_collected:>12,.2f}\n")
        
        if expenses:
            console.print("[bold]Expenses:[/bold]")
            for exp in expenses:
                console.print(f"  {exp['name']:<33} ${float(exp['total']):>12,.2f}")
            console.print("  " + "─" * 50)
        
        console.print(f"Total Expenses:        ${total_expenses:>12,.2f}\n")
        
        net_color = "green" if net_income >= 0 else "red"
        console.print(f"[{net_color}]Net Income:            ${net_income:>12,.2f}[/{net_color}]")
        
        if all:
            console.print("\n" + "─" * 60 + "\n")
    
    console.print()
