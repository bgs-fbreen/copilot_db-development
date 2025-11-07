"""
Project workbook management - Single XLSX file per project with:
  - Sheet 1: Baseline
  - Sheet 2+: Invoices (0001, 0002, etc.)
"""
import click
from rich.console import Console
from copilot.db import execute_query
from copilot.utils import get_project_directory_name
from datetime import datetime
from decimal import Decimal
import os
from pathlib import Path

console = Console()

# Project base directory
PROJECT_BASE_DIR = "/mnt/sda1/01_bgm_projman/Active"
PROJECT_FALLBACK_DIR = os.path.expanduser("~/bgm_projects/Active")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

def get_base_dir():
    """Get base project directory"""
    if os.path.exists(PROJECT_BASE_DIR):
        return PROJECT_BASE_DIR
    else:
        console.print(f"[yellow]⚠ Using fallback: {PROJECT_FALLBACK_DIR}[/yellow]")
        return PROJECT_FALLBACK_DIR

def create_project_directories(client_code, project_code, project_name=None):
    """
    Create complete project directory structure with abbreviated name
    """
    base_dir = get_base_dir()
    
    # Generate directory name with abbreviated project name
    dir_name = get_project_directory_name(project_code, project_name)
    
    console.print(f"[dim]Creating directory: {client_code}/{dir_name}[/dim]")
    
    # Project root directory
    project_root = os.path.join(base_dir, client_code, dir_name)
    
    # Subdirectories
    subdirs = [
        '01_baseline',
        '02_invoices',
        '03_authorization',
        '04_subcontractors',
        '05_reports'
    ]
    
    # Create all directories
    for subdir in subdirs:
        dir_path = os.path.join(project_root, subdir)
        Path(dir_path).mkdir(parents=True, exist_ok=True)
    
    return project_root

def get_project_workbook_path(client_code, project_code, project_name=None):
    """Get path to project workbook"""
    base_dir = get_base_dir()
    dir_name = get_project_directory_name(project_code, project_name)
    workbook_dir = os.path.join(base_dir, client_code, dir_name, '02_invoices')
    Path(workbook_dir).mkdir(parents=True, exist_ok=True)
    
    filename = f"{project_code}_invoices.xlsx"
    return os.path.join(workbook_dir, filename)

def get_or_create_workbook(client_code, project_code, project_name=None):
    """Get existing workbook or create new one with baseline sheet"""
    # Ensure all project directories exist
    create_project_directories(client_code, project_code, project_name)
    
    filepath = get_project_workbook_path(client_code, project_code, project_name)
    
    if os.path.exists(filepath):
        # Load existing workbook
        return load_workbook(filepath), filepath, False
    else:
        # Create new workbook with baseline
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb, filepath, True

def create_baseline_sheet(wb, project_code):
    """Create baseline sheet in workbook - USE res_id NOT res_name"""
    
    # Get baseline data
    data = get_baseline_data(project_code)
    if not data:
        return False
    
    proj = data['project']
    tasks = data['tasks']
    grand_total = data['grand_total']
    
    # Create baseline sheet
    ws = wb.create_sheet('Baseline', 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18  # Wider for subtotal label
    ws.column_dimensions['E'].width = 18  # Wider for totals
    ws.column_dimensions['F'].width = 50  # Wider for descriptions
    
    # Styles
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws[f'A{row}'] = "Baseline Costs"
    ws[f'A{row}'].font = title_font
    row += 1
    
    # Company header
    ws[f'A{row}'] = "Breen GeoScience Management, Inc."
    ws[f'A{row}'].font = header_font
    ws[f'E{row}'] = "Project"
    ws[f'E{row}'].font = header_font
    row += 1
    
    ws[f'E{row}'] = proj['project_name'] or proj['project_code']
    row += 1
    
    ws[f'A{row}'] = "Breen GeoScience Management, Inc."
    project_location = f"{proj['city']}, {proj['state']}" if proj['city'] else "USA"
    ws[f'E{row}'] = project_location
    row += 1
    
    ws[f'A{row}'] = "PMB #354, 4234 I-75 Business Spur"
    ws[f'E{row}'] = "USA"
    row += 1
    
    ws[f'A{row}'] = "Sault Ste. Marie, Michigan USA 49783"
    row += 2
    
    # Date and project info
    today = datetime.now()
    ws[f'E{row}'] = f"Date: {today.strftime('%m/%d/%y')}"
    ws[f'E{row}'].font = header_font
    row += 1
    
    # Client info
    ws[f'A{row}'] = "Client"
    ws[f'A{row}'].font = header_font
    ws[f'E{row}'] = f"BGM # {proj['project_code']}"
    row += 1
    
    ws[f'A{row}'] = proj['client_name']
    ws[f'E{row}'] = "A/R 45 days"
    row += 1
    
    # Client address
    if proj['street_address']:
        ws[f'A{row}'] = proj['street_address']
        row += 1
    
    if proj['city'] and proj['state']:
        ws[f'A{row}'] = f"{proj['city']}, {proj['state']} {proj['zip'] or ''}"
        row += 1
    
    ws[f'A{row}'] = "USA"
    row += 1
    
    if proj['contact_name']:
        ws[f'A{row}'] = f"Attention: {proj['contact_name']}"
        row += 1
    
    if proj['project_desc']:
        ws[f'A{row}'] = f"Project Description: {proj['project_desc']}"
        row += 2
    else:
        row += 2
    
    # Baseline items by task
    running_subtotal = Decimal('0')
    
    for task_key in sorted(tasks.keys()):
        task = tasks[task_key]
        
        # Task header
        ws[f'A{row}'] = f"{task['task_no']}: {task['task_name']}"
        ws[f'A{row}'].font = Font(bold=True, underline='single')
        row += 2
        
        # Column headers
        headers = ['Resource\nName', 'Units', 'Unit\nRate', 'Expense', 'Total', 'Description']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_font
            cell.border = border_thin
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True)
        row += 1
        
        # Resources - THIS IS WHERE res_id SHOULD APPEAR
        for res in task['resources']:
            ws[f'A{row}'] = res['resource']  # This should be res_id like "F.Breen"
            ws[f'B{row}'] = f"{res['units']:.2f}" if res['units'] else ""
            ws[f'C{row}'] = f"${res['rate']:.2f}" if res['rate'] else ""
            ws[f'D{row}'] = f"${res['expense']:,.2f}" if res['expense'] > 0 else ""
            ws[f'E{row}'] = f"${res['total']:,.2f}"
            ws[f'F{row}'] = res['description']
            
            # Format description cell - smaller font, wrap text
            ws[f'F{row}'].font = Font(size=9)
            ws[f'F{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border_thin
            
            # Auto-adjust row height for wrapped text (minimum 15, can grow)
            ws.row_dimensions[row].height = None  # Auto-height
            
            row += 1
        
        running_subtotal += task['subtotal']
        row += 1
    
    # Subtotal - right justified
    ws[f'D{row}'] = "Labor / Material / Expense Sub Total :"
    ws[f'D{row}'].font = bold_font
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'] = f"${running_subtotal:,.2f}"
    ws[f'E{row}'].font = bold_font
    row += 2
    
    # Grand total - right justified
    ws[f'D{row}'] = "Baseline Total :"
    ws[f'D{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'] = f"${grand_total:,.2f}"
    ws[f'E{row}'].font = Font(bold=True, size=12)
    row += 2
    
    # Footer
    ws[f'A{row}'] = today.strftime('%m/%d/%Y')
    ws[f'F{row}'] = "Page 1 of 1"
    
    return True

def get_baseline_data(project_code):
    """Fetch baseline data - USE res_id ONLY"""
    
    project = execute_query("""
        SELECT
            p.project_code,
            p.project_name,
            p.project_desc,
            p.client_po,
            c.code as client_code,
            c.name as client_name,
            c.contact_name,
            c.street_address,
            c.city,
            c.state,
            c.zip
        FROM bgs.project p
        JOIN bgs.client c ON c.code = p.client_code
        WHERE p.project_code = %s
    """, (project_code,))
    
    if not project:
        return None
    
    proj = project[0]
    
    # Get baseline - NO RESOURCE TABLE JOIN
    baseline = execute_query("""
        SELECT
            b.task_no,
            b.sub_task_no,
            t.task_name,
            b.res_id,
            b.base_units,
            b.base_rate,
            b.base_miles,
            b.base_miles_rate,
            b.base_expense,
            (COALESCE(b.base_units, 0) * COALESCE(b.base_rate, 0) +
             COALESCE(b.base_miles, 0) * COALESCE(b.base_miles_rate, 0) +
             COALESCE(b.base_expense, 0)) as total
        FROM bgs.baseline b
        LEFT JOIN bgs.task t ON 
            t.project_code = b.project_code AND
            t.task_no = b.task_no AND
            t.sub_task_no = b.sub_task_no
        WHERE b.project_code = %s
        ORDER BY b.task_no, b.sub_task_no, b.res_id
    """, (project_code,))
    
    tasks = {}
    grand_total = Decimal('0')
    
    for row in baseline:
        task_key = f"{row['task_no']}"
        task_name = row['task_name'] or "Consulting"
        
        if task_key not in tasks:
            tasks[task_key] = {
                'task_no': row['task_no'],
                'task_name': task_name,
                'resources': [],
                'subtotal': Decimal('0')
            }
        
        # USE res_id DIRECTLY - this is "F.Breen" not "Breen Geoscience"
        resource_display = row['res_id']
        console.print(f"[dim]Baseline resource: {resource_display}[/dim]")
        
        expense = Decimal(str(row['base_expense'] or 0))
        total = Decimal(str(row['total'] or 0))
        
        tasks[task_key]['resources'].append({
            'resource': resource_display,  # Should be "F.Breen"
            'units': row['base_units'],
            'rate': row['base_rate'],
            'expense': expense,
            'total': total,
            'description': task_name
        })
        
        tasks[task_key]['subtotal'] += total
        grand_total += total
    
    return {
        'project': proj,
        'tasks': tasks,
        'grand_total': grand_total
    }

def add_invoice_sheet(wb, invoice_code):
    """Add invoice sheet to workbook"""
    
    # Get invoice data
    data = get_invoice_data(invoice_code)
    if not data:
        return False
    
    inv = data['invoice']
    items = data['labor_items']
    totals = data['totals']
    
    # Create sheet with invoice number (e.g., "0001")
    sheet_name = f"{inv['invoice_number']:04d}"
    
    # Remove sheet if it already exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    ws = wb.create_sheet(sheet_name)
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 55  # Wider for descriptions
    
    # Styles
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws[f'A{row}'] = "Invoice Report"
    ws[f'A{row}'].font = title_font
    row += 1
    
    # Company header
    ws[f'A{row}'] = "Breen GeoScience Management, Inc."
    ws[f'A{row}'].font = header_font
    ws[f'F{row}'] = "Project"
    ws[f'F{row}'].font = header_font
    row += 1
    
    ws[f'A{row}'] = "Payment to:"
    ws[f'F{row}'] = inv['project_name'] or inv['project_code']
    row += 1
    
    ws[f'A{row}'] = "Breen GeoScience Management, Inc."
    project_location = f"{inv['client_city']}, {inv['client_state']}" if inv['client_city'] else "USA"
    ws[f'F{row}'] = project_location
    row += 1
    
    ws[f'A{row}'] = "PMB #354, 4234 I-75 Business Spur"
    ws[f'F{row}'] = "USA"
    row += 1
    
    ws[f'A{row}'] = "Sault Ste. Marie, Michigan USA 49783"
    row += 2
    
    ws[f'F{row}'] = f"Date: {inv['invoice_date'].strftime('%B %d, %Y')}"
    ws[f'F{row}'].font = header_font
    row += 1
    
    # Client section
    ws[f'A{row}'] = "Client"
    ws[f'A{row}'].font = header_font
    ws[f'F{row}'] = f"Desc. {inv['project_code']}"
    row += 1
    
    ws[f'A{row}'] = inv['client_name']
    ws[f'F{row}'] = f"PO#: {inv['client_po'] or 'NA'}"
    row += 1
    
    client_location = ""
    if inv['client_city'] and inv['client_state']:
        client_location = f"{inv['client_city']}, {inv['client_state']}"
    ws[f'A{row}'] = client_location
    ws[f'F{row}'] = inv['payment_terms'] or "A/R Net 30"
    row += 1
    
    ws[f'F{row}'] = f"Baseline: {inv['project_code']}"
    row += 1
    
    # Contact
    contact_line = ""
    if inv['contact_name']:
        if inv['contact_title']:
            contact_line = f"Attention: {inv['contact_name']}, {inv['contact_title']}"
        else:
            contact_line = f"Attention: {inv['contact_name']}"
    ws[f'A{row}'] = contact_line
    ws[f'F{row}'] = f"Invoice No. {inv['invoice_code']}"
    ws[f'F{row}'].font = header_font
    row += 1
    
    project_desc = f"Project Description: {inv['project_desc']}" if inv['project_desc'] else ""
    ws[f'A{row}'] = project_desc
    row += 2
    
    # Labor section
    ws[f'A{row}'] = "Labor"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline='single')
    row += 2
    
    # Column headers
    headers = ['Date', 'Resource Name', 'Units', 'Unit Rate', 'Task', 'Total', 'Description']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = border_thin
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    row += 1
    
    # Labor line items
    for item in items:
        ws[f'A{row}'] = item['date'].strftime('%Y-%m-%d')
        ws[f'B{row}'] = item['resource']  # This should be res_id like "F.Breen"
        ws[f'C{row}'] = f"{item['hours']:.2f}"
        ws[f'D{row}'] = f"${item['rate']:.2f}"
        ws[f'E{row}'] = item['task']
        ws[f'F{row}'] = f"${item['total']:.2f}"
        ws[f'G{row}'] = item['description']
        
        # Format description cell - smaller font, wrap text
        ws[f'G{row}'].font = Font(size=9)
        ws[f'G{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border_thin
        
        # Auto-adjust row height for wrapped text
        ws.row_dimensions[row].height = None  # Auto-height
        
        row += 1
    
    row += 1
    
    # Totals - right justified
    ws[f'A{row}'] = "Total Hrs."
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'C{row}'] = f"{totals['hours']:.2f}"
    ws[f'E{row}'] = "Total Labor"
    ws[f'E{row}'].font = bold_font
    ws[f'E{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'] = f"${totals['labor']:.2f}"
    ws[f'F{row}'].font = bold_font
    row += 2
    
    # Grand total
    ws[f'F{row}'] = f"${totals['grand_total']:.2f}"
    ws[f'F{row}'].font = Font(bold=True, size=14)
    
    return True

def get_invoice_data(invoice_code):
    """Fetch invoice data - USE res_id ONLY"""
    
    invoice = execute_query("""
        SELECT
            i.invoice_code,
            i.invoice_number,
            i.invoice_date,
            i.due_date,
            i.amount,
            i.paid_amount,
            i.status,
            i.payment_terms,
            p.project_code,
            p.project_name,
            p.project_desc,
            p.client_po,
            c.code as client_code,
            c.name as client_name,
            c.contact_name,
            c.contact_title,
            c.city as client_city,
            c.state as client_state
        FROM bgs.invoice i
        JOIN bgs.project p ON p.project_code = i.project_code
        JOIN bgs.client c ON c.code = p.client_code
        WHERE i.invoice_code = %s
    """, (invoice_code,))
    
    if not invoice:
        return None
    
    inv = invoice[0]
    
    # Get timesheets - NO RESOURCE TABLE JOIN
    timesheets = execute_query("""
        SELECT 
            t.ts_date,
            t.res_id,
            t.ts_units,
            t.task_no,
            t.ts_desc,
            b.base_rate
        FROM bgs.timesheet t
        LEFT JOIN bgs.baseline b ON 
            b.project_code = t.project_code AND
            b.task_no = t.task_no AND
            b.sub_task_no = t.sub_task_no AND
            b.res_id = t.res_id
        WHERE t.invoice_code = %s
        ORDER BY t.ts_date ASC, t.task_no ASC
    """, (invoice_code,))
    
    labor_items = []
    total_hours = Decimal('0')
    total_labor = Decimal('0')
    
    for ts in timesheets:
        hours = Decimal(str(ts['ts_units']))
        rate = Decimal(str(ts['base_rate'] or 0))
        labor_amt = hours * rate
        
        total_hours += hours
        total_labor += labor_amt
        
        # USE res_id DIRECTLY
        resource_display = ts['res_id']
        console.print(f"[dim]Invoice resource: {resource_display}[/dim]")
        
        labor_items.append({
            'date': ts['ts_date'],
            'resource': resource_display,  # Should be "F.Breen"
            'hours': hours,
            'rate': rate,
            'task': ts['task_no'],
            'total': labor_amt,
            'description': ts['ts_desc']
        })
    
    return {
        'invoice': inv,
        'labor_items': labor_items,
        'totals': {
            'hours': total_hours,
            'labor': total_labor,
            'grand_total': total_labor
        }
    }

@click.command('workbook')
@click.argument('project_code')
@click.option('--open', 'open_file', is_flag=True, help='Open workbook after creation')
def create_workbook(project_code, open_file):
    """Create or update project workbook with baseline"""
    
    if not XLSX_AVAILABLE:
        console.print("[red]XLSX export requires openpyxl: pip install openpyxl[/red]")
        return
    
    # Get project info
    proj = execute_query("""
        SELECT 
            p.project_code,
            p.project_name,
            p.client_code,
            c.name as client_name
        FROM bgs.project p
        JOIN bgs.client c ON c.code = p.client_code
        WHERE p.project_code = %s
    """, (project_code,))
    
    if not proj:
        console.print(f"[red]Project '{project_code}' not found[/red]")
        return
    
    p = proj[0]
    
    console.print(f"\n[bold cyan]Creating Project Workbook[/bold cyan]\n")
    console.print(f"[bold]Project:[/bold] {p['project_name']}")
    console.print(f"[bold]Client:[/bold] {p['client_name']}\n")
    
    # Get or create workbook
    wb, filepath, is_new = get_or_create_workbook(p['client_code'], p['project_code'], p['project_name'])
    
    # Create/update baseline sheet
    if create_baseline_sheet(wb, project_code):
        # Save workbook
        wb.save(filepath)
        
        console.print(f"[bold green]✓ Workbook created/updated![/bold green]")
        console.print(f"[dim]Location: {filepath}[/dim]")
        console.print(f"[dim]Sheet: Baseline[/dim]\n")
        
        if open_file:
            import subprocess
            subprocess.run(['xdg-open', filepath])
    else:
        console.print("[red]Error: No baseline data found[/red]")

@click.command('add-invoice-to-workbook')
@click.argument('invoice_code')
@click.option('--open', 'open_file', is_flag=True, help='Open workbook after adding invoice')
def add_invoice_to_workbook(invoice_code, open_file):
    """Add invoice sheet to project workbook"""
    
    if not XLSX_AVAILABLE:
        console.print("[red]XLSX export requires openpyxl: pip install openpyxl[/red]")
        return
    
    # Get invoice/project info
    inv_data = execute_query("""
        SELECT 
            i.invoice_code,
            i.invoice_number,
            p.project_code,
            p.project_name,
            p.client_code,
            c.name as client_name
        FROM bgs.invoice i
        JOIN bgs.project p ON p.project_code = i.project_code
        JOIN bgs.client c ON c.code = p.client_code
        WHERE i.invoice_code = %s
    """, (invoice_code,))
    
    if not inv_data:
        console.print(f"[red]Invoice '{invoice_code}' not found[/red]")
        return
    
    inv = inv_data[0]
    
    console.print(f"\n[bold cyan]Adding Invoice to Workbook[/bold cyan]\n")
    console.print(f"[bold]Invoice:[/bold] {inv['invoice_code']}")
    console.print(f"[bold]Project:[/bold] {inv['project_name']}\n")
    
    # Get or create workbook
    wb, filepath, is_new = get_or_create_workbook(inv['client_code'], inv['project_code'], inv['project_name'])
    
    # If new workbook, create baseline first
    if is_new:
        console.print("[yellow]Creating baseline sheet first...[/yellow]")
        create_baseline_sheet(wb, inv['project_code'])
    
    # Add invoice sheet
    if add_invoice_sheet(wb, invoice_code):
        # Save workbook
        wb.save(filepath)
        
        sheet_name = f"{inv['invoice_number']:04d}"
        console.print(f"[bold green]✓ Invoice added to workbook![/bold green]")
        console.print(f"[dim]Location: {filepath}[/dim]")
        console.print(f"[dim]Sheet: {sheet_name}[/dim]\n")
        
        if open_file:
            import subprocess
            subprocess.run(['xdg-open', filepath])
    else:
        console.print("[red]Error: Could not add invoice to workbook[/red]")

