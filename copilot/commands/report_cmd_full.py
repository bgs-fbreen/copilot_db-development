"""
Financial reports and analytics for Copilot Accounting System
"""
import click
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from copilot.db import execute_query, get_connection
from datetime import datetime, date, timedelta
from decimal import Decimal
import csv
import os
from pathlib import Path

console = Console()

def clear_screen():
    """Clear the terminal screen"""
    os.system('clear' if os.name != 'nt' else 'cls')

def get_report_dir():
    """Get or create the reports directory"""
    report_dir = os.path.expanduser("~/copilot_reports")
    Path(report_dir).mkdir(parents=True, exist_ok=True)
    return report_dir

def export_to_csv(data, filename, headers):
    """Export data to CSV file"""
    report_dir = get_report_dir()
    filepath = os.path.join(report_dir, filename)
    
    with open(filepath, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        for row in data:
            writer.writerow(row)
    
    return filepath

def export_to_xlsx(data, filename, headers, title):
    """Export data to XLSX file using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        report_dir = get_report_dir()
        filepath = os.path.join(report_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row_idx, row_data in enumerate(data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        return filepath
    except ImportError:
        console.print("[yellow]openpyxl not installed. Install with: pip install openpyxl[/yellow]")
        return None

@click.group()
def report():
    """Financial reports and analytics"""
    pass
