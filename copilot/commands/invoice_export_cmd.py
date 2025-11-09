"""
Invoice export and finalize command - creates/provides XLSX file for review, then finalizes only on request.
No references to invoice_item are present: all logic is based only on timesheet rows with invoice_code set.
"""

import click
from rich.console import Console
from rich.prompt import Confirm
from copilot.db import execute_query, get_connection
from decimal import Decimal
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle, numbers
import re

console = Console()

def sanitize(name):
    """Sanitize project_name for abbreviated directory naming."""
    if not name:
        return ""
    return re.sub(r'[^a-zA-Z0-9_]', '_', name.strip().replace(' ', '_'))[:32]

def create_project_dir(client_code, project_code, project_abbr):
    PROJECT_BASE = "/mnt/sda1/01_bgm_projman/Active"
    PROJECT_FALLBACK = os.path.expanduser("~/bgm_projects/Active")
    base_dir = PROJECT_BASE if os.path.exists(PROJECT_BASE) else PROJECT_FALLBACK
    project_root = os.path.join(base_dir, client_code, f"{project_code}_{project_abbr}")
    invoice_dir = os.path.join(project_root, '02_invoices')
    Path(invoice_dir).mkdir(parents=True, exist_ok=True)
    return invoice_dir

@click.command('export')
@click.argument('invoice_code')
def export_invoice(invoice_code):
    """
    Export a single invoice as a professionally styled XLSX.
    Gives the option to finalize the invoice (write invoice_code to timesheets, mark invoice as pending) only after review.
    No usage of invoice_item table: everything is based on timesheet rows with invoice_code set.
    """
    data = get_invoice_data(invoice_code)
    if not data:
        console.print(f"[red]Invoice '{invoice_code}' not found[/red]")
        return

    inv = data['invoice']
    project_abbr = sanitize(inv['project_name'] if inv['project_name'] else inv['project_code'])
    invoice_dir = create_project_dir(inv['client_code'], inv['project_code'], project_abbr)
    xlsx_file = os.path.join(invoice_dir, f"{invoice_code}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Invoice {inv['invoice_number']:04d}"

    set_invoice_header(ws, inv)

    current_row = 15  # Start after header

    # XLSX Number formatting
    currency_fmt = NamedStyle(name="currency_fmt", number_format=numbers.FORMAT_CURRENCY_USD_SIMPLE)
    int_fmt = NamedStyle(name="int_fmt", number_format='#,##0')

    if "currency_fmt" not in wb.named_styles:
        wb.add_named_style(currency_fmt)
    if "int_fmt" not in wb.named_styles:
        wb.add_named_style(int_fmt)

    styles = get_styles(wb)

    # Divide timesheets into labor (non-lump) and lump sum
    labor_items = [li for li in data['labor_items'] if not li['is_lump_sum']]
    lump_sum_items = [li for li in data['labor_items'] if li['is_lump_sum']]

    # --- Labor Section ---
    if labor_items:
        ws[f"A{current_row}"] = "Date"
        ws[f"B{current_row}"] = "Resource Name"
        ws[f"C{current_row}"] = "Units"
        ws[f"D{current_row}"] = "Unit Rate"
        ws[f"E{current_row}"] = "Task"
        ws[f"F{current_row}"] = "Total"
        ws[f"G{current_row}"] = "Description"
        for col in range(1, 8):
            ws[f"{chr(64+col)}{current_row}"].font = styles['header']
            ws[f"{chr(64+col)}{current_row}"].fill = styles['filled']
        current_row += 1
        for item in labor_items:
            ws[f"A{current_row}"] = item['date'].strftime('%Y-%m-%d')
            ws[f"B{current_row}"] = item['resource']
            ws[f"C{current_row}"] = float(item['hours'])
            ws[f"C{current_row}"].style = "int_fmt"
            ws[f"D{current_row}"] = float(item['rate']) if item['rate'] else None
            ws[f"D{current_row}"].style = "currency_fmt"
            ws[f"E{current_row}"] = item['task']
            ws[f"F{current_row}"] = float(item['total'])
            ws[f"F{current_row}"].style = "currency_fmt"
            ws[f"G{current_row}"] = item['description']
            current_row += 1
        # Totals - labor
        ws[f"A{current_row}"] = "Total Hrs."
        ws[f"C{current_row}"] = float(data['totals']['hours'])
        ws[f"C{current_row}"].style = "int_fmt"
        ws[f"E{current_row}"] = "Total Labor"
        ws[f"F{current_row}"] = float(data['totals']['labor'])
        ws[f"F{current_row}"].style = "currency_fmt"
        ws[f"A{current_row}"].font = styles['bold']
        ws[f"E{current_row}"].font = styles['bold']
        ws[f"F{current_row}"].font = styles['bold']
        current_row += 2

    # --- Lump Sum Section (if any) ---
    if lump_sum_items:
        ws[f"A{current_row}"] = "Lump Sum Tasks"
        ws[f"A{current_row}"].font = styles['section']
        current_row += 1
        ws[f"A{current_row}"] = "Date"
        ws[f"B{current_row}"] = "Resource Name"
        ws[f"E{current_row}"] = "Total"
        ws[f"G{current_row}"] = "Description"
        for col in (1,2,5,7):
            ws[f"{chr(64+col)}{current_row}"].font = styles['header']
            ws[f"{chr(64+col)}{current_row}"].fill = styles['filled']
        current_row += 1
        for item in lump_sum_items:
            ws[f"A{current_row}"] = item['date'].strftime('%Y-%m-%d')
            ws[f"B{current_row}"] = item['resource']
            ws[f"E{current_row}"] = float(item['total'])
            ws[f"E{current_row}"].style = "currency_fmt"
            ws[f"G{current_row}"] = item['description']
            current_row += 1
        # Lump sum subtotal
        ws[f"E{current_row}"] = "Subtotal:"
        ws[f"F{current_row}"] = float(data['totals']['lump_sum'])
        ws[f"F{current_row}"].style = "currency_fmt"
        ws[f"E{current_row}"].font = styles['bold']
        ws[f"F{current_row}"].font = styles['bold']
        current_row += 2

    # --- Mileage Section ---
    if data['totals']['mileage'] > 0:
        ws[f"E{current_row}"] = "Mileage Total:"
        ws[f"F{current_row}"] = float(data['totals']['mileage'])
        ws[f"F{current_row}"].style = "currency_fmt"
        ws[f"E{current_row}"].font = styles['bold']
        ws[f"F{current_row}"].font = styles['bold']
        current_row += 1

    # --- Expenses Section ---
    if data['totals']['expenses'] > 0:
        ws[f"E{current_row}"] = "Expenses Total:"
        ws[f"F{current_row}"] = float(data['totals']['expenses'])
        ws[f"F{current_row}"].style = "currency_fmt"
        ws[f"E{current_row}"].font = styles['bold']
        ws[f"F{current_row}"].font = styles['bold']
        current_row += 1

    # --- Grand Total ---
    ws[f"F{current_row}"] = data['totals']['grand_total']
    ws[f"F{current_row}"].style = "currency_fmt"
    ws[f"F{current_row}"].font = Font(bold=True, size=14)

    # Autofit columns
    for col in 'ABCDEFG':
        ws.column_dimensions[col].width = 17 if col != 'G' else 36

    wb.save(xlsx_file)
    console.print(f"[green]✓ Invoice exported: {xlsx_file}[/green]")
    console.print("[bold]This invoice is in DRAFT mode.[/bold] Review the XLSX file before finalizing.\n")

    if Confirm.ask("Do you want to finalize this invoice (writes to DB, locks timesheets)?", default=False):
        finalize_invoice(inv['project_code'], invoice_code)
        console.print(f"[green]✓ Invoice {invoice_code} finalized![/green]")
    else:
        console.print("[yellow]Invoice left as DRAFT, no database changes made.[/yellow]")

def finalize_invoice(project_code, invoice_code):
    """
    Update all null invoice_code in timesheet for this project to the given invoice_code,
    and update the invoice status in the invoice table (status -> 'pending').
    """
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            # Update only unlinked timesheets!
            cur.execute("""
                UPDATE bgs.timesheet
                SET invoice_code = %s
                WHERE project_code = %s AND invoice_code IS NULL
            """, (invoice_code, project_code))
            cur.execute("""
                UPDATE bgs.invoice
                SET status = 'pending'
                WHERE invoice_code = %s
            """, (invoice_code,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        console.print(f"[red]Error finalizing invoice: {e}[/red]")
    finally:
        conn.close()

def get_invoice_data(invoice_code):
    """
    Fetch invoice data (header and line items) from DB, using only timesheet for detail lines.
    """
    invoice = execute_query("""
        SELECT
            i.invoice_code, i.invoice_number, i.project_code, i.invoice_date, i.amount,
            i.status, i.due_date, i.notes,
            p.project_name, p.project_desc, p.client_po,
            c.code as client_code, c.name as client_name, c.contact_name,
            c.address, c.city, c.state, c.zip
        FROM bgs.invoice i
        JOIN bgs.project p ON p.project_code = i.project_code
        JOIN bgs.client c ON c.code = p.client_code
        WHERE i.invoice_code = %s
    """, (invoice_code,))

    if not invoice:
        return None

    inv = invoice[0]

    timesheets = execute_query("""
        SELECT 
            t.ts_date, t.res_id, COALESCE(r.res_name, t.res_id) as res_name, t.ts_units,
            t.task_no, t.sub_task_no, t.ts_desc, t.ts_mileage, t.ts_expense,
            b.base_rate, b.base_miles_rate, b.is_lump_sum
        FROM bgs.timesheet t
        LEFT JOIN bgs.resource r ON r.res_id = t.res_id
        LEFT JOIN bgs.baseline b ON 
            b.project_code = t.project_code
            AND b.task_no = t.task_no AND b.sub_task_no = t.sub_task_no AND b.res_id = t.res_id
        WHERE t.invoice_code = %s
        ORDER BY t.ts_date, t.task_no, t.sub_task_no
    """, (invoice_code,))

    labor_items = []
    total_hours = Decimal('0')
    total_labor = Decimal('0')
    total_mileage = Decimal('0')
    total_expenses = Decimal('0')
    lump_sum_total = Decimal('0')

    for ts in timesheets:
        is_lump_sum = bool(ts.get('is_lump_sum'))
        hours = Decimal(str(ts['ts_units'])) if ts['ts_units'] is not None else Decimal('0')
        rate = Decimal(str(ts['base_rate'] or 0))
        miles = Decimal(str(ts['ts_mileage'] or 0))
        mile_rate = Decimal(str(ts['base_miles_rate'] or 0))
        expense = Decimal(str(ts['ts_expense'] or 0))

        if is_lump_sum:
            total = expense
            lump_sum_total += total
        else:
            labor_amt = hours * rate
            mile_amt = miles * mile_rate
            total = labor_amt + mile_amt + expense
            total_hours += hours
            total_labor += labor_amt
            total_mileage += mile_amt
            total_expenses += expense

        task_key = f"{ts['task_no']}" + (f":{ts['sub_task_no']}" if ts['sub_task_no'] else "")

        labor_items.append({
            'date': ts['ts_date'],
            'resource': ts['res_name'],
            'hours': hours,
            'rate': rate,
            'task': task_key,
            'total': total,
            'miles': miles,
            'mile_rate': mile_rate,
            'mile_amt': miles * mile_rate,
            'expense': expense,
            'is_lump_sum': is_lump_sum,
            'description': ts['ts_desc'] or "",
        })

    grand_total = total_labor + total_mileage + total_expenses + lump_sum_total

    return {
        'invoice': inv,
        'labor_items': labor_items,
        'totals': {
            'hours': total_hours,
            'labor': total_labor,
            'mileage': total_mileage,
            'expenses': total_expenses,
            'lump_sum': lump_sum_total,
            'grand_total': grand_total
        }
    }

def set_invoice_header(ws, inv):
    """Set up the professional invoice header, matching baseline, with all info."""
    ws['A1'] = "Invoice Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = "Breen GeoScience Management, Inc."
    ws['A3'] = "Payment to:"
    ws['A4'] = "Breen GeoScience Management, Inc."
    ws['A5'] = "PMB #354, 4234 I-75 Business Spur"
    ws['A6'] = "Sault Ste. Marie, Michigan USA 49783"

    ws['F2'] = "Project"
    ws['F3'] = inv['project_name'] or ""
    ws['F4'] = f"{inv['city']}, {inv['state']}" if inv['city'] else ""
    ws['F5'] = "USA"
    ws['F7'] = f"Date: {inv['invoice_date'].strftime('%b. %d, %Y') if inv['invoice_date'] else ''}"
    ws['A8'] = "Client"
    ws['A9'] = inv['client_name']
    ws['A10'] = inv['address'] or ""
    ws['A11'] = f"{inv['city']}, {inv['state']}" if inv['city'] else ""
    ws['A12'] = "USA"
    ws['A13'] = f"Attention: {inv['contact_name']}" if inv['contact_name'] else ""
    ws['A14'] = f"Project Description: {inv['project_desc']}" if inv['project_desc'] else ""

    ws['F8'] = f"Desc. {inv['project_code']}"
    ws['F9'] = f"PO#: {inv['client_po'] or 'NA'}"
    ws['F10'] = "A/R Net 30"
    ws['F11'] = f"Baseline: {inv['project_code']}"
    ws['F13'] = f"Invoice No. {inv['invoice_code']}"

    # Style header fields
    for r in [2, 8]:
        ws[f'A{r}'].font = Font(bold=True)
    for r in [2, 3, 7, 8, 14]:
        ws[f'F{r}'].font = Font(bold=True)

def get_styles(wb):
    return {
        'header': Font(bold=True, size=11),
        'section': Font(bold=True, size=12, underline="single"),
        'filled': PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
        'bold': Font(bold=True),
    }
